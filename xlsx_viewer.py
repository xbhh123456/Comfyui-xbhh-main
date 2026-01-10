import os

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XBHHXlsxViewer:
    """
    XBHH XLSX 文档查看器节点
    
    功能:
    - 读取 Excel (.xlsx) 文件
    - 显示工作表内容
    - 输出格式化的表格文本
    """
    
    RETURN_TYPES = ("STRING", "INT", "INT")
    RETURN_NAMES = ("content", "row_count", "col_count")
    FUNCTION = "view_xlsx"
    CATEGORY = "XBHH"
    
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "file_path": ("STRING", {
                    "default": "",
                    "tooltip": "xlsx 文件的完整路径"
                }),
            },
            "optional": {
                "sheet_name": ("STRING", {
                    "default": "",
                    "tooltip": "工作表名称（留空则使用第一个工作表）"
                }),
                "max_rows": ("INT", {
                    "default": 100,
                    "min": 1,
                    "max": 10000,
                    "step": 1,
                    "tooltip": "最大显示行数"
                }),
            }
        }
    
    def view_xlsx(self, file_path, sheet_name="", max_rows=100):
        # 检查依赖
        if not OPENPYXL_AVAILABLE:
            return ("❌ 错误: 需要安装 openpyxl 库\n请运行: pip install openpyxl", 0, 0)
        
        # 检查文件
        if not file_path:
            return ("⚠️ 请输入文件路径", 0, 0)
        
        if not os.path.exists(file_path):
            return (f"❌ 文件不存在: {file_path}", 0, 0)
        
        if not file_path.lower().endswith('.xlsx'):
            return ("❌ 仅支持 .xlsx 格式文件", 0, 0)
        
        try:
            # 加载工作簿
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # 选择工作表
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title
            
            # 读取数据
            rows = []
            col_widths = []
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx > max_rows:
                    break
                
                # 转换为字符串
                row_data = [str(cell) if cell is not None else "" for cell in row]
                rows.append(row_data)
                
                # 计算列宽
                for col_idx, cell in enumerate(row_data):
                    cell_len = len(cell)
                    if col_idx >= len(col_widths):
                        col_widths.append(cell_len)
                    else:
                        col_widths[col_idx] = max(col_widths[col_idx], cell_len)
            
            wb.close()
            
            if not rows:
                return (f"📄 工作表 [{sheet_name}] 为空", 0, 0)
            
            # 格式化输出
            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0
            
            # 构建表格文本
            output_lines = []
            output_lines.append(f"📊 文件: {os.path.basename(file_path)}")
            output_lines.append(f"📋 工作表: {sheet_name}")
            output_lines.append(f"📏 大小: {total_rows} 行 × {total_cols} 列")
            if total_rows > max_rows:
                output_lines.append(f"⚠️ 仅显示前 {max_rows} 行")
            output_lines.append("-" * 50)
            
            # 限制列宽
            max_col_width = 20
            col_widths = [min(w, max_col_width) for w in col_widths]
            
            # 输出每一行
            for row_idx, row_data in enumerate(rows):
                formatted_cells = []
                for col_idx, cell in enumerate(row_data):
                    width = col_widths[col_idx] if col_idx < len(col_widths) else max_col_width
                    # 截断过长的内容
                    if len(cell) > width:
                        cell = cell[:width-2] + ".."
                    formatted_cells.append(cell.ljust(width))
                
                line = " | ".join(formatted_cells)
                output_lines.append(line)
                
                # 在第一行后添加分隔线（表头）
                if row_idx == 0:
                    separator = "-+-".join(["-" * w for w in col_widths])
                    output_lines.append(separator)
            
            content = "\n".join(output_lines)
            return (content, total_rows, total_cols)
            
        except Exception as e:
            return (f"❌ 读取失败: {str(e)}", 0, 0)


# 注册节点
NODE_CLASS_MAPPINGS = {
    "XBHHXlsxViewer": XBHHXlsxViewer
}
NODE_DISPLAY_NAME_MAPPINGS = {
    "XBHHXlsxViewer": "xbhh XLSX查看器"
}
